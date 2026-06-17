"""Excel import: downloadable template + parsing of an uploaded workbook."""
import io
from datetime import date, datetime

from dateutil import parser as dateparser
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy.orm import Session

from app import models

TASK_HEADERS = ["Name", "Responsible", "Offset days", "No deadline"]
ENTITY_HEADERS = [
    "Code",
    "Name",
    "Location",
    "Go-live",
    "On hold",
    "Next step",
    "Notes",
]

TRUE_WORDS = {"yes", "y", "true", "1", "igen", "x", "on hold", "hold"}


def _truthy(v) -> bool:
    if v is None:
        return False
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return v != 0
    return str(v).strip().lower() in TRUE_WORDS


def _to_int(v) -> int:
    if v is None or v == "":
        return 0
    try:
        return int(float(v))
    except (ValueError, TypeError):
        return 0


def _to_date(v):
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    try:
        return dateparser.parse(str(v), dayfirst=False).date()
    except (ValueError, TypeError, OverflowError):
        return None


# ---------------------------------------------------------------- template
def build_template() -> bytes:
    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F47F5")

    def style_header(ws, headers):
        for col, h in enumerate(headers, start=1):
            c = ws.cell(row=1, column=col, value=h)
            c.font = header_font
            c.fill = header_fill
        ws.freeze_panes = "A2"

    tasks = wb.active
    tasks.title = "Tasks"
    style_header(tasks, TASK_HEADERS)
    for row in [
        ["Backbone cabling complete", "Montana", 30, "no"],
        ["Migration complete / Go Live", "S&T", 0, "no"],
        ["Old equipment removed", "S&T", -1, "no"],
        ["Passive NW documentation", "Montana", 0, "yes"],
    ]:
        tasks.append(row)
    widths = [38, 16, 12, 12]
    for i, w in enumerate(widths, start=1):
        tasks.column_dimensions[tasks.cell(row=1, column=i).column_letter].width = w

    ents = wb.create_sheet("Entities")
    style_header(ents, ENTITY_HEADERS)
    for row in [
        [
            "001D",
            "Central Office Building",
            "https://maps.google.com/?q=47.28,18.89",
            "2026-07-08",
            "no",
            "",
            "",
        ],
        ["002A", "Site A", "", "", "no", "", ""],
    ]:
        ents.append(row)
    widths = [12, 28, 34, 14, 10, 26, 30]
    for i, w in enumerate(widths, start=1):
        ents.column_dimensions[ents.cell(row=1, column=i).column_letter].width = w

    info = wb.create_sheet("Read me")
    notes = [
        "How to use this template",
        "",
        "Tasks sheet = the repeating task template for the project.",
        "  Name        - task name (required)",
        "  Responsible - team or owner (optional)",
        "  Offset days - days relative to go-live. Positive = before go-live,",
        "                negative = after. 0 = on go-live day.",
        "  No deadline - yes/no. yes = never flagged overdue.",
        "",
        "Entities sheet = the items the tasks repeat over (racks, sites, ...).",
        "  Code      - short code (e.g. 001D). Used to match on re-import.",
        "  Name      - name / building (optional)",
        "  Location  - address or maps link (optional)",
        "  Go-live   - date, e.g. 2026-07-08. Leave empty if unknown.",
        "  On hold   - yes/no",
        "  Next step - free text (optional)",
        "  Notes     - free text (optional)",
        "",
        "Replace the example rows with your own data, then import the file.",
        "Sheet names (Tasks / Entities) and the header row must stay as they are.",
    ]
    for i, line in enumerate(notes, start=1):
        c = info.cell(row=i, column=1, value=line)
        if i == 1:
            c.font = Font(bold=True, size=13)
    info.column_dimensions["A"].width = 70

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------- parsing
def _find_sheet(wb, name, fallback_index):
    for s in wb.sheetnames:
        if s.strip().lower() == name.lower():
            return wb[s]
    if fallback_index < len(wb.sheetnames):
        return wb[wb.sheetnames[fallback_index]]
    return None


def _header_map(ws, expected):
    """Map expected header label -> column index based on the first row."""
    result = {}
    if ws is None:
        return result
    first = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not first:
        return result
    lowered = [str(h).strip().lower() if h is not None else "" for h in first]
    for label in expected:
        try:
            result[label] = lowered.index(label.lower())
        except ValueError:
            result[label] = None
    return result


def parse_workbook(data: bytes) -> dict:
    wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    tasks_ws = _find_sheet(wb, "Tasks", 0)
    ents_ws = _find_sheet(wb, "Entities", 1)

    tasks = []
    tmap = _header_map(tasks_ws, TASK_HEADERS)
    if tasks_ws is not None and tmap.get("Name") is not None:
        for row in tasks_ws.iter_rows(min_row=2, values_only=True):
            def g(label):
                idx = tmap.get(label)
                return row[idx] if idx is not None and idx < len(row) else None

            name = g("Name")
            if name is None or str(name).strip() == "":
                continue
            tasks.append(
                {
                    "name": str(name).strip(),
                    "responsible": str(g("Responsible") or "").strip(),
                    "offset_days": _to_int(g("Offset days")),
                    "no_deadline": _truthy(g("No deadline")),
                }
            )

    entities = []
    emap = _header_map(ents_ws, ENTITY_HEADERS)
    if ents_ws is not None and (
        emap.get("Code") is not None or emap.get("Name") is not None
    ):
        for row in ents_ws.iter_rows(min_row=2, values_only=True):
            def g(label):
                idx = emap.get(label)
                return row[idx] if idx is not None and idx < len(row) else None

            code = str(g("Code") or "").strip()
            name = str(g("Name") or "").strip()
            if not code and not name:
                continue
            entities.append(
                {
                    "code": code,
                    "name": name,
                    "location": str(g("Location") or "").strip(),
                    "golive_date": _to_date(g("Go-live")),
                    "on_hold": _truthy(g("On hold")),
                    "next_step": str(g("Next step") or "").strip(),
                    "notes": str(g("Notes") or "").strip(),
                }
            )
    wb.close()
    return {"tasks": tasks, "entities": entities}


# ---------------------------------------------------------------- apply
def apply_import(db: Session, project_id: int, parsed: dict, mode: str) -> dict:
    mode = (mode or "append").lower()
    counts = {
        "tasks_created": 0,
        "tasks_updated": 0,
        "entities_created": 0,
        "entities_updated": 0,
    }

    if mode == "replace":
        db.query(models.Entity).filter(
            models.Entity.project_id == project_id
        ).delete(synchronize_session=False)
        db.query(models.TaskDefinition).filter(
            models.TaskDefinition.project_id == project_id
        ).delete(synchronize_session=False)
        db.flush()

    # ---- tasks ----
    existing_defs = {
        d.name.strip().lower(): d
        for d in db.query(models.TaskDefinition)
        .filter(models.TaskDefinition.project_id == project_id)
        .all()
    }
    next_pos = (
        db.query(models.TaskDefinition)
        .filter(models.TaskDefinition.project_id == project_id)
        .count()
    )
    for t in parsed["tasks"]:
        key = t["name"].lower()
        d = existing_defs.get(key)
        if d:
            d.responsible = t["responsible"]
            d.offset_days = t["offset_days"]
            d.no_deadline = t["no_deadline"]
            counts["tasks_updated"] += 1
        else:
            d = models.TaskDefinition(
                project_id=project_id,
                name=t["name"],
                responsible=t["responsible"],
                offset_days=t["offset_days"],
                no_deadline=t["no_deadline"],
                position=next_pos,
            )
            db.add(d)
            existing_defs[key] = d
            next_pos += 1
            counts["tasks_created"] += 1
    db.flush()

    # ---- entities ----
    existing_ents = {
        e.code.strip().lower(): e
        for e in db.query(models.Entity)
        .filter(models.Entity.project_id == project_id)
        .all()
        if e.code
    }
    next_epos = (
        db.query(models.Entity)
        .filter(models.Entity.project_id == project_id)
        .count()
    )
    for e in parsed["entities"]:
        key = e["code"].lower() if e["code"] else None
        obj = existing_ents.get(key) if key else None
        if obj:
            obj.name = e["name"] or obj.name
            obj.location = e["location"] or obj.location
            obj.golive_date = e["golive_date"]
            obj.on_hold = e["on_hold"]
            obj.next_step = e["next_step"] or obj.next_step
            obj.notes = e["notes"] or obj.notes
            counts["entities_updated"] += 1
        else:
            obj = models.Entity(
                project_id=project_id,
                code=e["code"],
                name=e["name"],
                location=e["location"],
                golive_date=e["golive_date"],
                on_hold=e["on_hold"],
                next_step=e["next_step"],
                notes=e["notes"],
                position=next_epos,
            )
            db.add(obj)
            if key:
                existing_ents[key] = obj
            next_epos += 1
            counts["entities_created"] += 1
    db.flush()

    # ---- sync task instances for every entity x task definition ----
    all_defs = (
        db.query(models.TaskDefinition)
        .filter(models.TaskDefinition.project_id == project_id)
        .all()
    )
    all_ents = (
        db.query(models.Entity)
        .filter(models.Entity.project_id == project_id)
        .all()
    )
    def_ids = [d.id for d in all_defs]
    for ent in all_ents:
        existing = {
            i.task_def_id
            for i in db.query(models.TaskInstance)
            .filter(models.TaskInstance.entity_id == ent.id)
            .all()
        }
        for did in def_ids:
            if did not in existing:
                db.add(models.TaskInstance(entity_id=ent.id, task_def_id=did))
    db.commit()
    return counts
